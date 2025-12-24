import time
import platform
import os
import openpyxl
import win32com.client
from openpyxl.styles.borders import Border, Side
from openpyxl.styles.fonts import Font
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from openpyxl.worksheet.worksheet import Worksheet


class Excel:

    def __init__(self, closing_date)->None:
        self.__pf:str = platform.system()
        if self.__pf == 'Windows':
            path = r'//192.168.1.247/共有/営業課ﾌｫﾙﾀﾞ/02請求書/01Master/invoice_format.xlsx'
        else:
            path = r'/mnt/public/営業課ﾌｫﾙﾀﾞ/02請求書/01Master/invoice_format.xlsx'

        file_path = (path)
        self.wb:openpyxl.Workbook = openpyxl.load_workbook(filename=file_path,
                                                                 data_only=True)
        self.ws:Worksheet = self.wb['1']

        self.__closing_date:str = closing_date

    def filling_page_invoice(self, last_balance, deposit, carryover_price, 
                             sales_price, tax, billed_price, TAX_RATE 
                            )->None:
        '''
        1ページにしか記載されない情報
        '''
        self.ws.cell(row=22, column=1).value = last_balance
        self.ws.cell(row=22, column=7).value = deposit
        self.ws.cell(row=22, column=13).value = carryover_price
        self.ws.cell(row=22, column=18 ).value = sales_price
        self.ws.cell(row=22, column=23 ).value = tax
        self.ws.cell(row=22, column=28 ).value = sales_price + tax
        self.ws.cell(row=22, column=35 ).value = billed_price
        self.ws.cell(row=24, column=16 ).value = TAX_RATE + '%'
        self.ws.cell(row=24, column=18 ).value = sales_price
        self.ws.cell(row=24, column=23 ).value = tax


    def filling_page_customer(self, customer_code, customer_nam1, customer_nam2, 
                              customer_nam3, customer_pos1, customer_pos2,
                              customer_addr1, customer_addr2, customer_addr3,
                              page_count)->None:
        '''
        1ページの行数は70行、フォーマットは10ページ分ある
        '''
        self.ws.delete_rows(page_count * 70 + 1, 700)

        for i in range(page_count):
            self.ws.cell(row=i*70+1, column=40).value = page_count
            self.ws.cell(row=i*70+6, column=2).value = customer_code #type:ignore
            self.ws.cell(row=i*70+2, column=26).value = (self.__closing_date[:4]  + '年' + #type:ignore
                                         self.__closing_date[4:6] + '月' + 
                                         self.__closing_date[6:])
            self.ws.cell(row=i*70+1,column=2).value = ('〒' + customer_pos1 + '-' + 
                                                                 customer_pos2)
            self.ws.cell(row=i*70+2,column=2).value = customer_addr1
            self.ws.cell(row=i*70+3,column=2).value = customer_addr2
            self.ws.cell(row=i*70+4,column=2).value = customer_addr3
            self.ws.cell(row=i*70+7,column=2).value = customer_nam1
            self.ws.cell(row=i*70+8,column=2).value = customer_nam2 #type:ignore
            self.ws.cell(row=i*70+9,column=2).value = customer_nam3 #type:ignore
            if customer_nam2 == ' ' and customer_nam3 == ' ':
                self.ws.cell(row=i*70+7,column=19).value = '御中' #type:ignore
            if customer_nam2 != ' ' and customer_nam3 == ' ':
                self.ws.cell(row=i*70+8,column=19).value = '御中' #type:ignore
            if customer_nam3 != ' ':
                self.ws.cell(row=i*70+9,column=19).value = '御中' #type:ignore

    def filling_page_sales(self,row_no, count_page_sum_row, 
                           sales_deposits_count,
                           sale_no,
                           sale_date,
                           sale_hinban,
                           sale_hinmei,
                           sale_qty,
                           sale_tani,
                           sale_unit_price,
                           sale_price,
                           tekiyo,
                           toriKbn,
                           page_count)->None:

        last_page_sum_row = 0
        if page_count == 1:
            last_page_sum_row = 69
        last_page_sum_row = 139 + (page_count -2) * 70


        koumoku: str = ''
        if toriKbn == '01':
            koumoku = '売上'
        if toriKbn == '03':
            koumoku = '値引'
        if toriKbn == '02':
            koumoku = '返品'
        self.ws.cell(row=row_no, column=1).value = (sale_date[4:6] + '/' + 
                                                              sale_date[6:])
        self.ws.cell(row=row_no + 1, column=1).value = koumoku #type:ignore
        self.ws.cell(row=row_no, column=3).value = sale_no
        self.ws.cell(row=row_no, column=7).value = sale_hinban#type:ignore
        self.ws.cell(row=row_no + 1, column=7).value = sale_hinmei
        self.ws.cell(row=row_no, column=20).value = sale_qty
        self.ws.cell(row=row_no, column=24).value = sale_tani
        self.ws.cell(row=row_no, column=26).value = sale_unit_price
        self.ws.cell(row=row_no, column=30).value = sale_price
        self.ws.cell(row=row_no, column=34).value = tekiyo
        self.ws.cell(row=count_page_sum_row, column=26).value = ( #type:ignore
                    self.ws.cell(row=count_page_sum_row, column=26).value + 1)
        self.ws.cell(row=count_page_sum_row, column=29).value = (
            self.ws.cell(row=count_page_sum_row, column=29).value + sale_price)
        self.ws.cell(row=last_page_sum_row, column=26).value = (#type:ignore
                      self.ws.cell(row=last_page_sum_row, column=26).value + 1)
        self.ws.cell(row=last_page_sum_row+1, column=26).value = (#type:ignore
                    self.ws.cell(row=last_page_sum_row+1, column=26).value + 1)
        self.ws.cell(row=last_page_sum_row, column=29).value = (
             self.ws.cell(row=last_page_sum_row, column=29).value + sale_price)
        self.ws.cell(row=last_page_sum_row+1, column=29).value = (
           self.ws.cell(row=last_page_sum_row+1, column=29).value + sale_price)



    def filling_page_deposits(self,row_no, page_count, 
                              sales_deposits_count,
                              deposit_no,
                              deposit_date,
                              deposit_kubun,
                              deposit_price)->None:
        self.ws.cell(row=row_no, column=1).value = (deposit_date[4:6] + '/' + 
                                                             deposit_date[6:])
        self.ws.cell(row=row_no+1, column=1).value = deposit_kubun
        self.ws.cell(row=row_no, column=3).value = deposit_no
        self.ws.cell(row=row_no, column=7).value = '＜ 御 入 金 ＞' #type:ignore
        self.ws.cell(row=row_no, column=30).value = deposit_price

    def arrange_sheets(self, page_count)->None:
        for i in range(68, (page_count-1) * 70 + 69, 70):
            self.ws.cell(row=i, column=26).value = \
                            str(self.ws.cell(row=i, column=26).value) + '件' #type:ignore
            self.ws.cell(row=i + 1, column=26).value = \
                          str(self.ws.cell(row=i+1, column=26).value) + '件'#type:ignore
            self.ws.cell(row=i + 2, column=26).value = \
                          str(self.ws.cell(row=i+2, column=26).value) + '件'#type:ignore

        for i in range((page_count-1) * 70, 68, -70):
            self.ws.cell(row=i, column=21).value = None
            self.ws.cell(row=i, column=26).value = None
            self.ws.cell(row=i, column=29).value = None
            self.ws.cell(row=i-1, column=21).value = None
            self.ws.cell(row=i-1, column=26).value = None
            self.ws.cell(row=i-1, column=29).value = None

            sheet_range:object = self.ws.iter_rows(min_row=i-1, max_row=i, 
                                                           min_col=1, max_col=40)

            border_no:Border = Border(top=None, bottom=None, left=None, right=None)
            for row in sheet_range:
                for cell in row:
                    cell.border = border_no

        # 印貼り付け
        kakuin_path:str = r'//192.168.1.247/共有/営業課ﾌｫﾙﾀﾞ/02請求書/01Master/kakuin.png'
        matudoin_path:str = r'//192.168.1.247/共有/営業課ﾌｫﾙﾀﾞ/02請求書/01Master/matudoin.png'
        if self.__pf == 'Linux':
            kakuin_path = r'/mnt/public/営業課ﾌｫﾙﾀﾞ/02請求書/01Master/kakuin.png'
            matudoin_path = r'/mnt/public/営業課ﾌｫﾙﾀﾞ/02請求書/01Master/matudoin.png'
        kakuin_img:Image = Image(kakuin_path)
        matudoin_img:Image = Image(matudoin_path)
        kakuin_img.width = 70
        kakuin_img.height = 67
        matudoin_img.width = 28
        matudoin_img.height = 30
        self.ws.add_image(kakuin_img, 'AE3')
        self.ws.add_image(matudoin_img, 'AF16')
          
        # print設定
        CM:float = 1 / 2.54
        max_row:int = page_count * 70
        print_area:str = 'A1:AN' + str(max_row)
        self.ws.print_area = print_area 

        self.ws.page_margins.top = 0.7 * CM
        self.ws.page_margins.right = 1.0 * CM
        self.ws.page_margins.bottom = 0.2 * CM
        self.ws.page_margins.left = 1.0 * CM
        self.ws.page_margins.header = 0
        self.ws.page_margins.footer = 0
        



    def save_file(self, is_post, customer_code)->None:

        post_is_do_not_is:str = 'n'
        if is_post:
            post_is_do_not_is = 'y'

        new_folder = f'//192.168.1.247/共有/営業課ﾌｫﾙﾀﾞ/02請求書/02Excel/{self.__closing_date}'
        if self.__pf == 'Linux':
            new_folder:str = f'/mnt/public/営業課ﾌｫﾙﾀﾞ/02請求書/02Excel/{self.__closing_date}'
        os.makedirs(name=new_folder, exist_ok=True)
        self.wb.save(filename= f'{new_folder}/{self.__closing_date}_{customer_code}.xlsx')

        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False # Excelを見えなくする
        file = excel.Workbooks.Open(f'{new_folder}/{self.__closing_date}_{customer_code}.xlsx')
        #time.sleep(15)
        try:
            file.Worksheets("1").Select()
            pdf_folder = f'//192.168.1.247/共有/営業課ﾌｫﾙﾀﾞ/02請求書/03Pdf/{self.__closing_date}/01未提出'
            os.makedirs(name= pdf_folder, exist_ok= True)
            file.ActiveSheet.ExportAsFixedFormat(0, f'{pdf_folder}/{self.__closing_date}_{customer_code}_{post_is_do_not_is}_.pdf')
            submitted_folder = f'//192.168.1.247/共有/営業課ﾌｫﾙﾀﾞ/02請求書/03Pdf/{self.__closing_date}/02提出済'
            os.makedirs(name= submitted_folder, exist_ok= True)
        except Exception as e:
            print(e)
        finally:
            file.Close() # ワークブックを閉じる
            excel.Quit() # Excelアプリケーションプロセスを終了(タスクマネージャーからも消えるようにする)
            del file     # ヒープのメモリ解放(COMオブジェクトへの参照)
            del excel    # ヒープのメモリ解放(COMオブジェクトへの参照)



